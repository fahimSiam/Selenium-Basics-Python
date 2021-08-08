import openpyxl
from openpyxl import Workbook, load_workbook

path = "D:\\files\\codes\\vs\\selenium\\OpenPyXL\\Sample.xlsx"
wb=load_workbook(path)
ws=wb.active #CHOOSING THE POSTCODESHEET

#TASK1 READ DATA FROM xlsx
print('Data value in cell A2: ',ws['A2'].value)


#TASK2 2.write data in excel file.
ws['A4']='4545'
print(ws['A4'].value)


#TASK 3. read all column data.
total_col=ws.max_column+1
print('printing all the values on column 1')
for rows in range (1,total_col):
    cell = ws.cell(column = 1, row = rows)
    print(cell.value)

#TASK 4. write  5 row data
empty_row=1
for cell in ws["C"]: #loop to ditermine empty row
    if cell.value is None:
        empty_row=cell.row
        break
print('empty row: ',empty_row)
values = [['6545','342', 'siam', '2014', '9999.12344'],
         ['2541','456', 'sami', '2015', '23454.675'], 
         ['4345','567', 'madeia', '2016', '45436.6565'], 
         ['6786','456', 'mangha', '2017', '946547.54564'],
         ['2345','764', 'bristy', '2018', '934534534.3453']]
for row, row_entries in enumerate(values, start=empty_row):
        for column, value in enumerate(row_entries, start=1):
            ws.cell(column=column, row=row, value=value)
wb.save(filename='Sample.xlsx')