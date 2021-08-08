import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
wb=load_workbook('Sample.xlsx')
ws=wb.active
cell = ws.cell(row = 1, column = 1) 
ws['A4']='7676'
print(ws['A4'].value)
print(wb.sheetnames)
wb.save(filename='Sample.xlsx')

#PRINTING WHOLE DATA
for row in range(1,10):
    s=""
    for col in range(1,6):
        char=get_column_letter(col)
        s+=(str(ws[char+str(row)].value))
        s+=" "
    print(s)


#MERGING CELLS
ws.merge_cells("A2:A3")
ws.unmerge_cells("A2:A3")



#INSERT
ws.insert_cols(3)
ws.delete_cols(3)

ws['B2'].font= Font(bold=True, color="0099CCFF")

wb.save(filename='Sample.xlsx')